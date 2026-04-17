"""Rebuild CR-Aware Update xlsx from the LATEST 99-ticket original.

Key changes vs v1:
  - Base file is the user's updated 99-ticket original (incl. PAYM-176, PAYM-277)
  - All original rows are PRESERVED in place — nothing overwritten structurally
  - CR-aware additions are APPENDED after each tab's existing content
  - Only targeted cells are updated: title, dates, review method, classifications
    for reclassified tickets, and the new column G in All Tickets
"""
import shutil
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

DOCS = Path("/Users/guybaxter/Documents/Claude Code/hireup/docs")
SRC = DOCS / "Hireup Scope Analysis — Finalised.xlsx"
DST = DOCS / "Hireup Scope Analysis — Finalised (CR-Aware Update).xlsx"

assert SRC.exists()
shutil.copy(SRC, DST)

# ---------- Style palette (matches original exactly) ----------
TITLE = Font(name="Calibri", size=14, bold=True, color="000000")
SUBTITLE = Font(name="Calibri", size=10, color="000000")
SECTION = Font(name="Calibri", size=12, bold=True, color="073763")
TABLE_HDR = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
BODY = Font(name="Calibri", size=10, color="000000")
META_LABEL = Font(name="Calibri", size=11, bold=True, color="000000")

FILL_SECTION = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
FILL_HEADER = PatternFill(start_color="073763", end_color="073763", fill_type="solid")
FILL_IN_SCOPE = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
FILL_CREEP = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
FILL_BORDER = PatternFill(start_color="FCE5CD", end_color="FCE5CD", fill_type="solid")
FILL_CR_RAISED = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
FILL_CR_NEW = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
FILL_CR_COVERED = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
FILL_NONE = PatternFill(fill_type=None)

THIN = Side(style="thin", color="BFBFBF")
BORDER_T = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
WRAP_CENTER = Alignment(wrap_text=True, vertical="center")


def banner(ws, row, text, span=4):
    c = ws.cell(row=row, column=1, value=text)
    c.font = SECTION
    c.fill = FILL_SECTION
    c.alignment = WRAP_CENTER
    c.border = BORDER_T
    try:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    except Exception:
        pass


def hdr(ws, row, values):
    for i, v in enumerate(values, 1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = TABLE_HDR
        c.fill = FILL_HEADER
        c.alignment = WRAP_CENTER
        c.border = BORDER_T


def body(cell, fill=None):
    cell.font = BODY
    cell.alignment = WRAP
    cell.border = BORDER_T
    if fill is not None:
        cell.fill = fill


wb = openpyxl.load_workbook(DST)

# ============================================================================
# Tab 1 — Executive Summary: append CR-aware sections at the bottom
# ============================================================================
ws = wb["1. Executive Summary"]

# Update title (row 1) and dates/method in-place
ws["A1"] = "Hireup JIRA Scope Analysis — Finalised Report (CR-Aware Update, 17 April 2026)"
ws["A1"].font = TITLE
ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

# Update Analysis Date and Review Method (find rows)
for r in range(1, 15):
    label = str(ws.cell(row=r, column=1).value or "")
    if label == "Analysis Date":
        ws.cell(row=r, column=2, value="17 April 2026")
        ws.cell(row=r, column=2).font = BODY
    elif label == "Review Method":
        ws.cell(row=r, column=2, value="Multi-agent independent review (4 passes — CR-aware, post CR01/CR02/Project Evolution discovery)")
        ws.cell(row=r, column=2).font = BODY

# Find the end of existing content
last_r = ws.max_row
start = last_r + 3

# ---- CR EXPOSURE — CR-AWARE RECALCULATION ----
banner(ws, start, "COMMERCIAL EXPOSURE — CR-AWARE RECALCULATION (17/04/26)")
hdr(ws, start + 1, ["", "Low Estimate", "High Estimate", "% of SOW"])
exposure = [
    ("CR01 Shift-Level Payroll Tax (raised 27/11/25)", 16000, 16000, "9%", FILL_CR_RAISED),
    ("CR02 Pay Advice (raised 20/1/26 — rewrite recommended)", 6000, 6000, "4%", FILL_CR_RAISED),
    ("CR02-Uplift (realistic T&M delta — recommended)", 9000, 16500, "5-10%", FILL_CR_NEW),
    ("CR-1 Sleepover Engine (not raised, 7 tickets)", 20000, 25000, "12-15%", None),
    ("CR-2 Portable LSL (not raised, 3 tickets)", 25000, 35000, "15-21%", None),
    ("CR-3 Enhancements & Integrations (not raised, 9 tickets)", 14000, 20000, "8-12%", None),
    ("CR-4 Project Evolution Addendum — NEW RECOMMENDED (Stages 4-5)", 31000, 50000, "18-29%", FILL_CR_NEW),
    ("Goodwill Concessions", -13000, -20000, "Relationship investment", None),
    ("Net Recoverable Total", 108000, 142500, "64-84%", FILL_HEADER),
]
for i, (lbl, lo, hi, pct, fill) in enumerate(exposure, start=start + 2):
    row_cells = [
        ws.cell(row=i, column=1, value=lbl),
        ws.cell(row=i, column=2, value=lo),
        ws.cell(row=i, column=3, value=hi),
        ws.cell(row=i, column=4, value=pct),
    ]
    for c in row_cells:
        body(c, fill)
        if lbl == "Net Recoverable Total":
            c.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    row_cells[1].number_format = '"$"#,##0;[Red]-"$"#,##0'
    row_cells[2].number_format = '"$"#,##0;[Red]-"$"#,##0'

start_cr = start + 2 + len(exposure) + 2

# ---- RECOMMENDED CR PACKAGING (CR-AWARE) ----
banner(ws, start_cr, "RECOMMENDED CR PACKAGING (7 CRs — CR-aware)")
hdr(ws, start_cr + 1, ["CR", "Description", "Est. Value", "Priority & Status"])
packaging = [
    ("CR01", "Shift-Level Payroll Tax & FP/NFP Cost Item", "$16,000 flat", "P0 — RAISED 27/11/25, issue covering letter before signing", FILL_CR_RAISED),
    ("CR02", "Additional Information in Pay Advices", "$6,000 T&M", "P0 — RAISED 20/1/26, REWRITE before signing", FILL_CR_RAISED),
    ("CR02-Uplift", "CR02 rewrite to realistic T&M", "+$9k-$16.5k", "P1 — bundle with CR02 rewrite", FILL_CR_NEW),
    ("CR-1", "Sleepover Engine Extension (7 tickets)", "$20k-$25k", "P0 — raise before PPT2 completes", None),
    ("CR-2", "Portable LSL Compliance (3 tickets)", "$25k-$35k", "P0 — raise before PPT2, zero ambiguity", None),
    ("CR-3", "Enhancements & Integrations bundle (9 tickets)", "$14k-$20k", "P1 — after CR-1/CR-2", None),
    ("CR-4", "Project Evolution — Month-End & Audit Trail (NEW)", "$31k-$50k", "P0 — NEW, raise before any Stage 4/5 dev", FILL_CR_NEW),
    ("", "Package discount (all remaining CRs)", "10-15% off", "Incentivise single approval", None),
]
for i, (cr, desc, est, pri, fill) in enumerate(packaging, start=start_cr + 2):
    cells = [
        ws.cell(row=i, column=1, value=cr),
        ws.cell(row=i, column=2, value=desc),
        ws.cell(row=i, column=3, value=est),
        ws.cell(row=i, column=4, value=pri),
    ]
    for c in cells:
        body(c, fill)

start_kf = start_cr + 2 + len(packaging) + 2

# ---- KEY FINDINGS (CR-AWARE) ----
banner(ws, start_kf, "KEY FINDINGS (CR-aware)")
hdr(ws, start_kf + 1, ["#", "Finding", "Commercial Impact", ""])
findings = [
    (1, "CR01 Shift-Level Payroll Tax is NET-NEW scope (no existing ticket coverage). $16k materially under-priced — true effort $24k–$30k.", "Sign CR01 but carve out PE Stages 4-5 before countersignature"),
    (2, "CR02 'significantly more complex than initially identified from Discovery' language admits 2c9 fault (SOW 3.2). Weaponisable by Hireup against CR-1/CR-3.", "REWRITE CR02 this week — reframe as 'emergent requirement during build'"),
    (3, "Project Evolution (Dec 2025) Stages 4-5 (month-end recalc, SRO Part/All toggle, reversal/correction/linkage) are unpriced scope creep beyond CR01.", "NEW CR-4 Project Evolution Addendum — $31k-$50k"),
    (4, "Sleepover Engine (7 tickets) — SOW Gap 2 narrow. Min top-ups, 18+hr, back-to-back, OT aggregation, midday penalties all beyond.", "CR-1 — $20k-$25k (raise before PPT2)"),
    (5, "Portable LSL (3 tickets) — not in SOW, RW, DW, or RAID. Highest defensibility.", "CR-2 — $25k-$35k (raise before PPT2)"),
    (6, "Hibiscus Release (PAYM-262) — SOW §5 Assumption 8 exclusion. Same logic supports excluding PE Stages 4-5.", "CR-3 / CR-4 contractual anchor"),
    (7, "Simplification Cluster (PAYM-327-330) — Rework of initial designs. SOW commits to one build cycle.", "Negotiate 50/50 cost share"),
    (8, "Custom UI Widget (PAYM-334) — Already built. No recovery mechanism beyond commercial agreement.", "Include in CR-3 at 50-60% rate"),
    (9, "Min Contracted Hours (PAYM-271) — JIRA comments pre-acknowledge 'most likely a change request'.", "CR-3 — client pre-concedes"),
    (10, "Revised net recoverable: $108k-$142.5k (64-84% of $170k SOW) vs original $46k-$67k (27-39%).", "+$62k-$75.5k uplift contingent on actions #1-3 this week"),
]
for i, (num, finding, impact) in enumerate(findings, start=start_kf + 2):
    cells = [
        ws.cell(row=i, column=1, value=num),
        ws.cell(row=i, column=2, value=finding),
        ws.cell(row=i, column=3, value=impact),
    ]
    for c in cells:
        body(c)

start_ta = start_kf + 2 + len(findings) + 2

# ---- TOP ACTIONS (CR-AWARE) ----
banner(ws, start_ta, "TOP ACTIONS (CR-aware, prioritised)")
hdr(ws, start_ta + 1, ["#", "Action", "Timing", ""])
actions = [
    (1, "Withdraw CR02 draft and rewrite at 60-90 hrs ($15k-$22.5k); strip 'Discovery under-scoped' phrasing", "This week", FILL_CR_NEW),
    (2, "Issue CR01 covering letter carving out month-end recalc + PAYM-98 GL report + SRO Part/All toggle", "Before CR01 countersignature", FILL_CR_NEW),
    (3, "Raise CR-4 Project Evolution Addendum (Stages 4-5)", "Within 2 weeks, before any dev", None),
    (4, "Raise CR-1 Sleepover Engine Extension ($20k-$25k)", "Within 2 weeks — before PPT2", None),
    (5, "Raise CR-2 Portable LSL Compliance ($25k-$35k)", "Within 2 weeks — before PPT2", None),
    (6, "Raise CR-3 Enhancements Bundle; PAYM-334 at 50-60% rate", "After CR-1/CR-2 lodged", None),
    (7, "Document in writing that CR01's 24hr T&M did NOT cover correction/linkage or month-end recalc", "This week", FILL_CR_NEW),
    (8, "Apply goodwill concessions: 50/50 on PAYM-327/328/330; absorb PAYM-347, 335", "With CR-1/CR-2 lodgement", None),
    (9, "Reclassify: PAYM-167, 329, 337, 342 → 'Covered by CR02'; PAYM-358 → 'Covered by CR01'", "Within 1 week", None),
    (10, "Sequence: CR02-rewrite + CR01-letter (now) → CR-1/CR-2 (2 wks) → CR-4 (2-4 wks) → CR-3 (post-PPT2)", "Ongoing", None),
]
for i, (num, action, timing, fill) in enumerate(actions, start=start_ta + 2):
    cells = [
        ws.cell(row=i, column=1, value=num),
        ws.cell(row=i, column=2, value=action),
        ws.cell(row=i, column=3, value=timing),
    ]
    for c in cells:
        body(c, fill)

# ---- Column widths — keep original layout intact ----
# (original widths preserved from copy)

# ============================================================================
# Tab 2 — All Tickets: add Column G "CR Coverage" without disturbing others
# ============================================================================
ws2 = wb["2. All Tickets"]

# Insert new column after Classification (col F) -> col G
ws2.insert_cols(7)

# Header
hdr_cell = ws2.cell(row=1, column=7, value="CR Coverage / Status")
hdr_cell.font = TABLE_HDR
hdr_cell.fill = FILL_HEADER
hdr_cell.alignment = WRAP_CENTER
hdr_cell.border = BORDER_T

# Ticket → (new classification label if changed, coverage note)
ticket_updates = {
    "PAYM-45":  (None, "→ CR-1 Sleepover Engine", None),
    "PAYM-98":  (None, "→ CR-4 (report build) + CR01 (data model)", FILL_CR_NEW),
    "PAYM-115": (None, "CR01 complexity uplift; may need CR-4 topup", None),
    "PAYM-121": (None, "Original PE CR1 (pre-Nov 2025 — unchanged)", None),
    "PAYM-123": (None, "CR02 adjacent (pay-code consolidation touch)", None),
    "PAYM-145": (None, "Split — CR01 state logic + CR-2 PLSL linkage", None),
    "PAYM-151": (None, "→ CR-3 (push-to-CR)", None),
    "PAYM-159": (None, "→ CR-2 Portable LSL", None),
    "PAYM-166": (None, "→ CR-3 Enhancements", None),
    "PAYM-167": ("In Scope → Covered by CR02", "COVERED BY CR02", FILL_CR_COVERED),
    "PAYM-173": (None, "→ CR-2 Portable LSL", None),
    "PAYM-232": (None, "→ CR-3 or Absorb", None),
    "PAYM-241": (None, "→ CR-2 Portable LSL", None),
    "PAYM-243": (None, "Original PE CR1 (pre-Nov 2025 — unchanged)", None),
    "PAYM-262": (None, "→ CR-3 Enhancements", None),
    "PAYM-271": (None, "→ CR-3 Enhancements", None),
    "PAYM-272": (None, "→ CR-1 (permanent fix)", None),
    "PAYM-273": (None, "→ CR-1 (midnight-crossing expansion)", None),
    "PAYM-279": (None, "→ CR-3 (FDVL standalone; NOT covered by CR02)", None),
    "PAYM-280": (None, "→ CR-3 Enhancements", None),
    "PAYM-293": (None, "Split — bug absorb; CAG/PPL → CR-3", None),
    "PAYM-311": (None, "→ CR-3", None),
    "PAYM-312": (None, "→ CR-1 Sleepover Engine", None),
    "PAYM-313": (None, "→ CR-1 Sleepover Engine", None),
    "PAYM-319": (None, "→ CR-1 Sleepover Engine", None),
    "PAYM-324": (None, "→ CR-1 Sleepover Engine", None),
    "PAYM-325": (None, "→ CR-1 Sleepover Engine", None),
    "PAYM-326": (None, "→ CR-1 Sleepover Engine", None),
    "PAYM-327": (None, "Goodwill 50/50 (Simplification)", FILL_BORDER),
    "PAYM-328": (None, "Goodwill 50/50 (Simplification)", FILL_BORDER),
    "PAYM-329": ("Borderline → Covered by CR02", "COVERED BY CR02", FILL_CR_COVERED),
    "PAYM-330": (None, "Goodwill 50/50 (Simplification)", FILL_BORDER),
    "PAYM-333": (None, "→ CR-3 Enhancements", None),
    "PAYM-334": (None, "→ CR-3 @ 50-60% rate (already built)", None),
    "PAYM-335": (None, "Absorb (likely config defect)", FILL_BORDER),
    "PAYM-337": ("Borderline → Covered by CR02", "COVERED BY CR02 (email wrapper)", FILL_CR_COVERED),
    "PAYM-342": ("Borderline → Covered by CR02", "COVERED BY CR02 + CR01 reversal", FILL_CR_COVERED),
    "PAYM-347": (None, "Absorb (goodwill, trivial)", FILL_BORDER),
    "PAYM-354": (None, "→ CR-3 Enhancements", None),
    "PAYM-357": (None, "CR01 QA uplift (dual cost items)", None),
    "PAYM-358": ("Borderline → Covered by CR01", "COVERED BY CR01 (dual-cost-item template)", FILL_CR_COVERED),
}

# Apply updates
for r in range(2, ws2.max_row + 1):
    tk = ws2.cell(row=r, column=1).value
    if not tk:
        continue

    upd = ticket_updates.get(tk)
    # Column G default — empty white cell with body styling + border
    cov_cell = ws2.cell(row=r, column=7, value="")
    cov_cell.font = BODY
    cov_cell.alignment = WRAP
    cov_cell.border = BORDER_T

    if upd:
        new_class, coverage, cov_fill = upd
        # Update classification if changed
        if new_class:
            cls = ws2.cell(row=r, column=6, value=new_class)
            cls.font = BODY
            cls.alignment = WRAP
            cls.border = BORDER_T
            cls.fill = FILL_CR_COVERED
        # Write coverage
        cov_cell.value = coverage
        if cov_fill:
            cov_cell.fill = cov_fill
        elif coverage.startswith("→"):
            cov_cell.fill = FILL_CR_RAISED
        elif "Goodwill" in coverage or "Absorb" in coverage:
            cov_cell.fill = FILL_BORDER

# Column G width
ws2.column_dimensions["G"].width = 42

# Re-apply auto filter + freeze
ws2.freeze_panes = "A2"
ws2.auto_filter.ref = ws2.dimensions

# ============================================================================
# Tab 3 — Scope Creep — CR Themes: append CR01/CR02/CR-4 at end (before NOTES)
# ============================================================================
ws3 = wb["3. Scope Creep — CR Themes"]

# Find NOTES row
notes_row = None
for r in range(1, ws3.max_row + 1):
    if str(ws3.cell(row=r, column=1).value or "").strip() == "NOTES":
        notes_row = r
        break

# Append 3 new themes AFTER the existing last theme row, before NOTES
# New rows go at notes_row, shifting NOTES + below down by 3
if notes_row:
    ws3.insert_rows(notes_row, amount=3)

new_themes = [
    ("CR01 — Shift-Level Payroll Tax & FP/NFP (RAISED 27/11/25)",
     "Net-new — no existing ticket coverage", 0, "5", "L",
     "$16,000 flat (under-priced — true $24k-$30k)",
     "Dec-2025 Project Evolution programme; shift-level payroll tax, FP/NFP cost items not in SOW, RW, or DW",
     "Structural payroll tax changes: TSE Location field, paired FP/NFP cost items per state, payroll engine selection, journal reversal capability",
     "None — signed CR",
     FILL_CR_RAISED),
    ("CR02 — Additional Information in Pay Advices (RAISED 20/1/26)",
     "PAYM-167, 329, 337, 342 (+123 adjacent)", 4, "4", "M",
     "$6,000 T&M (under-priced — true $15k-$22.5k); REWRITE RECOMMENDED",
     "Gap 7 rescope + net-new shift-level data sourcing + layout redesign. CR language admits Discovery under-scoped (SOW 3.2 risk).",
     "Custom Pay Advice: consolidated rate, shift-level detail, redesigned layout",
     "'You under-scoped Discovery' (own-mouth admission) — must be softened before signing",
     FILL_CR_RAISED),
    ("CR-4 Project Evolution Addendum (Stages 4-5) — NEW RECOMMENDED",
     "Month-end recalc, SRO toggle, reversal/correction/linkage, retrospective UAT", 0, "5", "L",
     "$31,000-$50,000",
     "Dec-2025 Project Evolution doc introduces Stages 4-5 beyond CR01. Not in SOW, RW-PE, CR01, or RAID.",
     "Month-end recalculation engine with FP-only vs Part/All mode; SRO ruling response; three-way audit trail (original → reversal → correction)",
     "'CR01 already covered journal reversal' — rebut with CR01 24hr T&M never covered correction/linkage or month-end recalc",
     FILL_CR_NEW),
]

for i, theme in enumerate(new_themes, start=notes_row):
    vals = theme[:-1]
    fill = theme[-1]
    for c, val in enumerate(vals, start=1):
        cell = ws3.cell(row=i, column=c, value=val)
        body(cell, fill)

ws3.auto_filter.ref = ws3.dimensions

# ============================================================================
# Tab 4 — Borderline: update Commercial Approach column for reclassified tickets
# ============================================================================
ws4 = wb["4. Borderline"]

# Find header row (Ticket column)
hrow = None
for r in range(1, 10):
    if str(ws4.cell(row=r, column=1).value or "").strip() == "Ticket":
        hrow = r
        break

approach_updates = {
    "PAYM-151": ("→ CR-3 (push-to-CR)", None),
    "PAYM-232": ("→ CR-3 or Absorb", None),
    "PAYM-272": ("→ CR-1 (permanent fix)", None),
    "PAYM-273": ("→ CR-1 (midnight-crossing expansion)", None),
    "PAYM-293": ("Split — bug absorb; CAG/PPL → CR-3", None),
    "PAYM-311": ("→ CR-3", None),
    "PAYM-327": ("Goodwill 50/50 (Simplification)", FILL_BORDER),
    "PAYM-328": ("Goodwill 50/50 (Simplification)", FILL_BORDER),
    "PAYM-329": ("COVERED BY CR02 (consolidation)", FILL_CR_COVERED),
    "PAYM-330": ("Goodwill 50/50 (Simplification)", FILL_BORDER),
    "PAYM-335": ("Absorb (likely config defect)", FILL_BORDER),
    "PAYM-337": ("COVERED BY CR02 (email wrapper)", FILL_CR_COVERED),
    "PAYM-342": ("COVERED BY CR02 + CR01 reversal", FILL_CR_COVERED),
    "PAYM-347": ("Absorb (goodwill, trivial)", FILL_BORDER),
    "PAYM-358": ("COVERED BY CR01 (dual-cost-item template)", FILL_CR_COVERED),
}

if hrow:
    for r in range(hrow + 1, ws4.max_row + 1):
        tk = ws4.cell(row=r, column=1).value
        if tk in approach_updates:
            text, fill = approach_updates[tk]
            cell = ws4.cell(row=r, column=6, value=text)
            body(cell, fill)

# Update Resolution Buckets if present
for r in range(1, ws4.max_row + 1):
    val = str(ws4.cell(row=r, column=1).value or "")
    if val == "Likely Absorb":
        ws4.cell(row=r, column=2, value="PAYM-272, 273, 335, 347 (PAYM-329 now covered by CR02)")
        ws4.cell(row=r, column=3, value=4)
    elif val == "Negotiate Shared Cost":
        ws4.cell(row=r, column=2, value="PAYM-232, 293, 327, 328, 330 (PAYM-342→CR02, PAYM-358→CR01)")
        ws4.cell(row=r, column=3, value=5)
    elif val == "Push to CR":
        ws4.cell(row=r, column=2, value="PAYM-151, 311 (PAYM-337 now covered by CR02)")
        ws4.cell(row=r, column=3, value=2)

ws4.auto_filter.ref = ws4.dimensions

# ============================================================================
# Tab 5 — Review Findings: append new methodology + CR-aware strategy
# ============================================================================
ws5 = wb["5. Review Findings"]

# Update subtitle on row 2
ws5["A2"] = "Independent multi-agent review — 16 April 2026; CR-aware update — 17 April 2026"
ws5["A2"].font = SUBTITLE

# Insert methodology steps 5-6 after step 4
step4_row = None
for r in range(1, ws5.max_row + 1):
    if str(ws5.cell(row=r, column=1).value or "").startswith("4."):
        step4_row = r
        break

if step4_row:
    ws5.insert_rows(step4_row + 1, amount=2)
    for i, (lbl, desc) in enumerate([
        ("5. CR Discovery (17/04/26)",
         "Three new documents surfaced: CR01 Shift-Level Payroll Tax (raised 27/11/25, $16k); CR02 Additional Info in Pay Advices (raised 20/1/26, $6k); Project Evolution process doc (Dec 2025). 4-agent review pass."),
        ("6. Commercial Reconciliation",
         "Revised recoverable from $46k-$67k to $108k-$142.5k (64-84% of SOW). Identified CR-4 Project Evolution Addendum ($31k-$50k) as net-new scope. Reclassified 5 tickets as 'Covered by CR'.")
    ]):
        r = step4_row + 1 + i
        a = ws5.cell(row=r, column=1, value=lbl)
        b = ws5.cell(row=r, column=2, value=desc)
        body(a, FILL_CR_NEW)
        body(b, FILL_CR_NEW)

# Append CR-Aware Negotiation Strategy section
end = ws5.max_row + 3
banner(ws5, end, "CR-AWARE NEGOTIATION STRATEGY (17/04/26)")
hdr(ws5, end + 1, ["#", "Recommendation", "", ""])
new_strategies = [
    (1, "REWRITE CR02 before signing — remove 'more complex than initially identified from Discovery' phrasing. Discovery is 2c9 deliverable (SOW 3.2). Reframe as 'emergent requirement during build'."),
    (2, "COVERING LETTER on CR01 before countersignature — carve out: (i) month-end recalculation mode, (ii) PAYM-98 GL report build, (iii) SRO Part/All toggle."),
    (3, "Raise CR-4 Project Evolution Addendum BEFORE any Stage 4/5 development begins. Document Dec-2025 PE doc post-dates Nov-2025 CR01 as 'emergent post-CR'."),
    (4, "Maintain existing CR-1/CR-2/CR-3 stack — CR01/CR02 do NOT reduce those positions. CR01 is net-new; CR02 covers 5 tickets but the other 14 scope creep + 10 borderline remain."),
    (5, "Timing: CR02-rewrite + CR01-covering-letter THIS WEEK → CR-1/CR-2 (2 wks, before PPT2) → CR-4 (2-4 wks) → CR-3 (post-PPT2)."),
    (6, "Precedent risk: if Hireup accept CR02's 'Discovery under-scoped' framing, reusable for CR-1/CR-3. If rejected, damages CR-1/CR-3. REMOVE the phrase to protect the stack."),
    (7, "Cite SOW §5 Assumption 8 in CR-4 covering note — same logic that excludes Hibiscus supports excluding PE Stages 4-5."),
    (8, "Revised target recoverable: $108k-$142.5k (64-84% of $170k SOW). Contingent on (a) CR02 rewrite, (b) CR01 carve-out letter, (c) CR-4 raised pre-dev."),
]
for i, (num, rec) in enumerate(new_strategies, start=end + 2):
    a = ws5.cell(row=i, column=1, value=num)
    b = ws5.cell(row=i, column=2, value=rec)
    body(a)
    body(b)

wb.save(DST)
print(f"WROTE {DST}")
