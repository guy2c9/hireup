"""Build parallel Finalised xlsx with CR-aware updates — mirrors original format.

Strategy:
  1. Copy original xlsx (preserves all formatting, widths, colours, freeze panes, filters).
  2. Open the copy and patch cells in place with CR01/CR02/Project-Evolution findings.
  3. Save as a new file. Do not touch the original.
"""
import shutil
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from copy import copy

DOCS = Path("/Users/guybaxter/Documents/Claude Code/hireup/docs")
SRC = DOCS / "Hireup Scope Analysis — Finalised.xlsx"
DST = DOCS / "Hireup Scope Analysis — Finalised (CR-Aware Update).xlsx"

assert SRC.exists(), f"source missing: {SRC}"
shutil.copy(SRC, DST)

wb = openpyxl.load_workbook(DST)

# ---------- Fills (mirroring original palette) ----------
SCOPE_CREEP = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
BORDERLINE = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
IN_SCOPE = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
CR_COVERED = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")  # blue — CR covered
CR_NEW = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")  # orange — NEW/recommended
CR_RAISED = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # yellow — raised
HEADER_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
SUB_FONT = Font(name="Calibri", size=11, bold=True, color="1F3864")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")

# ===========================================================================
# Tab 1 — Executive Summary
# ===========================================================================
ws = wb["1. Executive Summary"]

# Title
ws["A1"] = "Hireup JIRA Scope Analysis — Finalised Report (CR-Aware Update, 17 April 2026)"

# Metadata
ws["B7"] = "17 April 2026"
ws["B9"] = "Multi-agent independent review (4 passes — CR-aware)"

# Commercial Exposure Summary — rewrite rows 17–23
ws["A17"] = ""
ws["B17"] = "Low Estimate"
ws["C17"] = "High Estimate"
ws["D17"] = "% of SOW"

exposure = [
    ("CR01 Shift-Level Payroll Tax (raised 27/11/25)", 16000, 16000, "9%"),
    ("CR02 Pay Advice (raised 20/1/26 — rewrite recommended)", 6000, 6000, "4%"),
    ("CR02-Uplift (recommended rewrite delta)", 9000, 16500, "5-10%"),
    ("CR-1 Sleepover Engine (not raised, 7 tickets)", 20000, 25000, "12-15%"),
    ("CR-2 Portable LSL (not raised, 3 tickets)", 25000, 35000, "15-21%"),
    ("CR-3 Enhancements & Integrations (not raised, 9 tickets)", 14000, 20000, "8-12%"),
    ("CR-4 Project Evolution Addendum — NEW RECOMMENDED (Stages 4-5)", 31000, 50000, "18-29%"),
    ("Goodwill Concessions", -13000, -20000, "Relationship investment"),
    ("Net Recoverable Total", 108000, 142500, "64-84%"),
]
for i, (label, lo, hi, pct) in enumerate(exposure, start=18):
    ws.cell(row=i, column=1, value=label)
    c_lo = ws.cell(row=i, column=2, value=lo)
    c_hi = ws.cell(row=i, column=3, value=hi)
    ws.cell(row=i, column=4, value=pct)
    c_lo.number_format = '"$"#,##0;[Red]-"$"#,##0'
    c_hi.number_format = '"$"#,##0;[Red]-"$"#,##0'
    # shade
    status = str(label)
    for col in range(1, 5):
        cell = ws.cell(row=i, column=col)
        cell.alignment = WRAP
        cell.border = BORDER
        if "raised 27/11" in status or "raised 20/1" in status:
            cell.fill = CR_RAISED
        elif "NEW RECOMMENDED" in status or "CR02-Uplift" in status:
            cell.fill = CR_NEW
        elif "Net Recoverable" in status:
            cell.fill = HEADER_FILL
            cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

# Clear any leftover rows from original block
for r in range(27, 30):
    for c in range(1, 6):
        ws.cell(row=r, column=c, value=None)
        ws.cell(row=r, column=c).fill = PatternFill(fill_type=None)

# CR Packaging section — shift to row 28
ws["A28"] = "RECOMMENDED CR PACKAGING (7 CRs — post-CR-aware analysis)"
ws["A28"].font = SUB_FONT
ws["A29"] = "CR"
ws["B29"] = "Description"
ws["C29"] = "Est. Value"
ws["D29"] = "Priority & Status"
for c in range(1, 5):
    cell = ws.cell(row=29, column=c)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(wrap_text=True, vertical="center")
    cell.border = BORDER

cr_packaging = [
    ("CR01", "Shift-Level Payroll Tax & FP/NFP Cost Item", "$16,000 flat", "P0 — RAISED 27/11/25, issue covering letter before signing"),
    ("CR02", "Additional Information in Pay Advices", "$6,000 T&M", "P0 — RAISED 20/1/26, REWRITE before signing (remove Discovery admission)"),
    ("CR02-Uplift", "CR02 rewrite to realistic T&M", "+$9k-$16.5k", "P1 — bundle with CR02 rewrite"),
    ("CR-1", "Sleepover Engine Extension (7 tickets)", "$20k-$25k", "P0 — raise before PPT2 completes"),
    ("CR-2", "Portable LSL Compliance (3 tickets)", "$25k-$35k", "P0 — raise before PPT2, zero ambiguity"),
    ("CR-3", "Enhancements & Integrations bundle (9 tickets)", "$14k-$20k", "P1 — after CR-1/CR-2"),
    ("CR-4", "Project Evolution — Month-End & Audit Trail (NEW)", "$31k-$50k", "P0 — NEW, raise before any Stage 4/5 dev"),
    ("", "Package discount (all remaining CRs)", "10-15% off", "Incentivise single approval"),
]
for i, row in enumerate(cr_packaging, start=30):
    for c, val in enumerate(row, start=1):
        cell = ws.cell(row=i, column=c, value=val)
        cell.alignment = WRAP
        cell.border = BORDER
        if row[0] in ("CR01", "CR02"):
            cell.fill = CR_RAISED
        elif row[0] in ("CR02-Uplift", "CR-4"):
            cell.fill = CR_NEW

# Key findings — add new ones at existing section
# Original rows 30-37 — shift and rewrite
ws["A39"] = "KEY FINDINGS (CR-aware)"
ws["A39"].font = SUB_FONT
ws["A40"] = "#"
ws["B40"] = "Finding"
ws["C40"] = "Commercial Impact"
for c in range(1, 4):
    cell = ws.cell(row=40, column=c)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(wrap_text=True, vertical="center")
    cell.border = BORDER

findings = [
    (1, "CR01 Shift-Level Payroll Tax is NET-NEW scope (no existing ticket coverage). $16k materially under-priced — true effort $24k–$30k. Existing 3-CR pipeline ($46k–$67k) remains fully live.", "Sign CR01 but carve out PE Stages 4-5 before countersignature"),
    (2, "CR02 'significantly more complex than initially identified from Discovery' language admits 2c9 fault (SOW 3.2 makes Discovery a 2c9 deliverable). Weaponisable by Hireup's counsel against CR-1/CR-3.", "REWRITE CR02 this week — reframe as 'emergent requirement during build'"),
    (3, "Project Evolution (Dec 2025) Stages 4-5 (month-end recalc, SRO Part/All toggle, reversal→correction→linkage audit trail) are unpriced scope creep beyond CR01.", "NEW CR-4 Project Evolution Addendum — $31k-$50k"),
    (4, "Sleepover Engine Extension (7 tickets) — SOW Gap 2 narrow wording. Minimum top-ups, 18+hr, back-to-back, OT aggregation, midday penalties all beyond.", "CR-1 — $20k-$25k (raise before PPT2)"),
    (5, "Portable LSL (3 tickets) — not in SOW, RW, DW, or RAID. Highest defensibility.", "CR-2 — $25k-$35k (raise before PPT2)"),
    (6, "Hibiscus Release (PAYM-262) — SOW §5 Assumption 8 explicit exclusion. Same logic supports excluding PE Stages 4-5.", "CR-3 / CR-4 contractual anchor"),
    (7, "Simplification Cluster (PAYM-327-330) — Rework of initial designs. SOW commits to one build cycle. Shared accountability.", "Negotiate 50/50 cost share"),
    (8, "Custom UI Widget (PAYM-334) — Already built. No recovery mechanism beyond commercial agreement.", "Include in CR-3 at 50-60% rate"),
    (9, "Min Contracted Hours (PAYM-271) — JIRA comments pre-acknowledge 'most likely a change request'.", "CR-3 — client pre-concedes"),
    (10, "Revised net recoverable: $108k-$142.5k (64-84% of $170k SOW) vs original $46k-$67k (27-39%).", "+$62k-$75.5k uplift contingent on actions #1-3 this week"),
]

for i, (num, finding, impact) in enumerate(findings, start=41):
    ws.cell(row=i, column=1, value=num)
    ws.cell(row=i, column=2, value=finding)
    ws.cell(row=i, column=3, value=impact)
    for c in range(1, 4):
        cell = ws.cell(row=i, column=c)
        cell.alignment = WRAP
        cell.border = BORDER

# Top Actions — rewrite
action_start = 41 + len(findings) + 2
ws.cell(row=action_start, column=1, value="TOP ACTIONS (CR-aware, prioritised)").font = SUB_FONT
ws.cell(row=action_start + 1, column=1, value="#")
ws.cell(row=action_start + 1, column=2, value="Action")
ws.cell(row=action_start + 1, column=3, value="Timing")
for c in range(1, 4):
    cell = ws.cell(row=action_start + 1, column=c)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(wrap_text=True, vertical="center")
    cell.border = BORDER

actions = [
    (1, "Withdraw CR02 draft and rewrite at 60-90 hrs ($15k-$22.5k); strip 'Discovery under-scoped' phrasing", "This week"),
    (2, "Issue CR01 covering letter carving out month-end recalc + PAYM-98 GL report + SRO Part/All toggle", "Before CR01 countersignature"),
    (3, "Raise CR-4 Project Evolution Addendum (Stages 4-5)", "Within 2 weeks, before any dev"),
    (4, "Raise CR-1 Sleepover Engine Extension ($20k-$25k)", "Within 2 weeks — before PPT2"),
    (5, "Raise CR-2 Portable LSL Compliance ($25k-$35k)", "Within 2 weeks — before PPT2"),
    (6, "Raise CR-3 Enhancements Bundle; PAYM-334 at 50-60% rate", "After CR-1/CR-2 lodged"),
    (7, "Document in writing that CR01's 24hr T&M did NOT cover correction/linkage or month-end recalc", "This week"),
    (8, "Apply goodwill concessions: 50/50 on PAYM-327/328/330; absorb PAYM-347, 335", "With CR-1/CR-2 lodgement"),
    (9, "Reclassify: PAYM-167, 329, 337, 342 → 'Covered by CR02'; PAYM-358 → 'Covered by CR01'", "Within 1 week"),
    (10, "Sequence: CR02-rewrite + CR01-letter (now) → CR-1/CR-2 (2 wks) → CR-4 (2-4 wks) → CR-3 (post-PPT2)", "Ongoing"),
]
for i, (num, action, timing) in enumerate(actions, start=action_start + 2):
    ws.cell(row=i, column=1, value=num)
    ws.cell(row=i, column=2, value=action)
    ws.cell(row=i, column=3, value=timing)
    for c in range(1, 4):
        cell = ws.cell(row=i, column=c)
        cell.alignment = WRAP
        cell.border = BORDER
        if "This week" in str(timing):
            cell.fill = CR_NEW

# Widen column A for summary labels
ws.column_dimensions["A"].width = 58
ws.column_dimensions["B"].width = 20
ws.column_dimensions["C"].width = 18
ws.column_dimensions["D"].width = 30

# ===========================================================================
# Tab 2 — All Tickets (add CR Coverage column)
# ===========================================================================
ws2 = wb["2. All Tickets"]

# Insert new column G after F (Classification)
# openpyxl insert_cols default shifts right
ws2.insert_cols(7)
ws2["G1"] = "CR Coverage / Status"
ws2["G1"].fill = HEADER_FILL
ws2["G1"].font = HEADER_FONT
ws2["G1"].alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

# Build ticket map
ticket_updates = {
    "PAYM-45":  ("Scope Creep", "→ CR-1 Sleepover Engine"),
    "PAYM-98":  ("Scope Creep", "→ CR-4 (report build) + CR01 (data model)"),
    "PAYM-115": ("In Scope", "CR01 complexity uplift; may need CR-4 topup"),
    "PAYM-121": ("In Scope", "Original PE CR1 (pre-Nov 2025 — unchanged)"),
    "PAYM-123": ("In Scope", "CR02 adjacent (pay-code consolidation touch)"),
    "PAYM-145": ("In Scope", "Split — CR01 state logic + CR-2 PLSL linkage"),
    "PAYM-151": ("Borderline", "→ CR-3 (push-to-CR)"),
    "PAYM-159": ("Scope Creep", "→ CR-2 Portable LSL"),
    "PAYM-166": ("Scope Creep", "→ CR-3 Enhancements"),
    "PAYM-167": ("In Scope → Covered by CR02", "COVERED BY CR02"),
    "PAYM-173": ("Scope Creep", "→ CR-2 Portable LSL"),
    "PAYM-232": ("Borderline", "→ CR-3 or Absorb"),
    "PAYM-241": ("Scope Creep", "→ CR-2 Portable LSL"),
    "PAYM-243": ("In Scope", "Original PE CR1 (pre-Nov 2025 — unchanged)"),
    "PAYM-262": ("Scope Creep", "→ CR-3 Enhancements"),
    "PAYM-271": ("Scope Creep", "→ CR-3 Enhancements"),
    "PAYM-272": ("Borderline", "→ CR-1 (permanent fix)"),
    "PAYM-273": ("Borderline", "→ CR-1 (midnight-crossing expansion)"),
    "PAYM-279": ("Scope Creep", "→ CR-3 (FDVL standalone, NOT covered by CR02)"),
    "PAYM-280": ("Scope Creep", "→ CR-3 Enhancements"),
    "PAYM-293": ("Borderline", "Split — bug absorb; CAG/PPL → CR-3"),
    "PAYM-311": ("Borderline", "→ CR-3"),
    "PAYM-312": ("Scope Creep", "→ CR-1 Sleepover Engine"),
    "PAYM-313": ("Scope Creep", "→ CR-1 Sleepover Engine"),
    "PAYM-319": ("Scope Creep", "→ CR-1 Sleepover Engine"),
    "PAYM-324": ("Scope Creep", "→ CR-1 Sleepover Engine"),
    "PAYM-325": ("Scope Creep", "→ CR-1 Sleepover Engine"),
    "PAYM-326": ("Scope Creep", "→ CR-1 Sleepover Engine"),
    "PAYM-327": ("Borderline", "Goodwill 50/50 (Simplification)"),
    "PAYM-328": ("Borderline", "Goodwill 50/50 (Simplification)"),
    "PAYM-329": ("Borderline → Covered by CR02", "COVERED BY CR02"),
    "PAYM-330": ("Borderline", "Goodwill 50/50 (Simplification)"),
    "PAYM-333": ("Scope Creep", "→ CR-3 Enhancements"),
    "PAYM-334": ("Scope Creep", "→ CR-3 @ 50-60% rate (already built)"),
    "PAYM-335": ("Borderline", "Absorb (likely config defect)"),
    "PAYM-337": ("Borderline → Covered by CR02", "COVERED BY CR02 (email wrapper)"),
    "PAYM-342": ("Borderline → Covered by CR02", "COVERED BY CR02 + CR01 reversal mechanism"),
    "PAYM-347": ("Borderline", "Absorb (goodwill, trivial)"),
    "PAYM-354": ("Scope Creep", "→ CR-3 Enhancements"),
    "PAYM-357": ("In Scope", "CR01 QA uplift (dual cost items)"),
    "PAYM-358": ("Borderline → Covered by CR01", "COVERED BY CR01 (dual-cost-item template)"),
}

max_row = ws2.max_row
for row in range(2, max_row + 1):
    ticket = ws2.cell(row=row, column=1).value
    if not ticket:
        continue
    update = ticket_updates.get(ticket)
    if update:
        new_class, coverage = update
        ws2.cell(row=row, column=6, value=new_class)
        cell_cov = ws2.cell(row=row, column=7, value=coverage)
        cell_cov.alignment = WRAP
        cell_cov.border = BORDER
        if "COVERED BY" in coverage:
            cell_cov.fill = CR_COVERED
        elif "→ CR-4" in coverage or "NEW" in coverage:
            cell_cov.fill = CR_NEW
        elif "Goodwill" in coverage or "Absorb" in coverage:
            cell_cov.fill = BORDERLINE
        # reapply classification colour if changed
        cls_cell = ws2.cell(row=row, column=6)
        cls_cell.alignment = WRAP
        cls_cell.border = BORDER
        if "Covered by" in new_class:
            cls_cell.fill = CR_COVERED
        elif new_class == "Scope Creep":
            cls_cell.fill = SCOPE_CREEP
        elif new_class == "Borderline":
            cls_cell.fill = BORDERLINE
        elif new_class == "In Scope":
            cls_cell.fill = IN_SCOPE
    else:
        ws2.cell(row=row, column=7, value="")

# Resize CR Coverage column
ws2.column_dimensions["G"].width = 48
# Re-apply filter
ws2.auto_filter.ref = ws2.dimensions

# ===========================================================================
# Tab 3 — Scope Creep — CR Themes (append CR01, CR02, CR-4 rows)
# ===========================================================================
ws3 = wb["3. Scope Creep — CR Themes"]

# Find last data row before NOTES section
max_r = ws3.max_row
notes_row = None
for r in range(1, max_r + 1):
    if str(ws3.cell(row=r, column=1).value or "").strip() == "NOTES":
        notes_row = r
        break

if notes_row is None:
    notes_row = max_r + 1

# Shift NOTES block down to make room for 3 new rows
shift = 4
for r in range(max_r, notes_row - 1, -1):
    for c in range(1, 10):
        src = ws3.cell(row=r, column=c)
        dst = ws3.cell(row=r + shift, column=c)
        dst.value = src.value
        if src.has_style:
            dst.font = copy(src.font)
            dst.fill = copy(src.fill)
            dst.border = copy(src.border)
            dst.alignment = copy(src.alignment)
            dst.number_format = src.number_format
        src.value = None
        src.fill = PatternFill(fill_type=None)

new_theme_start = notes_row
new_themes = [
    ("CR01 — Shift-Level Payroll Tax & FP/NFP (RAISED 27/11/25)", "Net-new — no existing ticket coverage", 0, "5", "L", "$16,000 flat (under-priced — true $24k-$30k)",
     "Dec-2025 Project Evolution programme; shift-level payroll tax, FP/NFP cost items not in SOW, RW, or DW",
     "Structural payroll tax changes: TSE Location field, paired FP/NFP cost items per state, payroll engine selection, journal reversal capability",
     "None — signed CR"),
    ("CR02 — Additional Information in Pay Advices (RAISED 20/1/26)", "PAYM-167, 329, 337, 342 (+PAYM-123 adjacent)", 4, "4", "M", "$6,000 T&M (under-priced — true $15k-$22.5k); REWRITE RECOMMENDED",
     "Gap 7 rescope + net-new shift-level data sourcing + layout redesign. CR language admits Discovery under-scoped (SOW 3.2 risk).",
     "Custom Pay Advice: consolidated rate, shift-level detail, redesigned layout",
     "'You under-scoped Discovery' (own-mouth admission) — must be softened before signing"),
    ("CR-4 Project Evolution Addendum (Stages 4-5) — NEW RECOMMENDED", "Month-end recalc engine, SRO toggle, reversal/correction/linkage, retrospective UAT", 0, "5", "L", "$31,000-$50,000",
     "Dec-2025 Project Evolution doc introduces Stages 4-5 beyond CR01. Not in SOW, RW-PE, CR01, or RAID. Emerged post-CR01 sign-off.",
     "Month-end recalculation engine with FP-only vs Part/All mode; SRO ruling response; three-way audit trail (original → reversal → correction)",
     "'CR01 already covered journal reversal' — rebut with CR01 24hr T&M never covered correction/linkage or month-end recalc"),
]

for i, theme in enumerate(new_themes, start=new_theme_start):
    for c, val in enumerate(theme, start=1):
        cell = ws3.cell(row=i, column=c, value=val)
        cell.alignment = WRAP
        cell.border = BORDER
        if "RAISED" in str(theme[0]):
            cell.fill = CR_RAISED
        elif "NEW RECOMMENDED" in str(theme[0]):
            cell.fill = CR_NEW

# Update NOTES totals at bottom
# Find new notes row
new_notes_start = notes_row + shift
for r in range(new_notes_start, ws3.max_row + 1):
    val = ws3.cell(row=r, column=1).value or ""
    if "Total estimated scope creep value" in str(val):
        ws3.cell(row=r, column=1, value="6. Total estimated recoverable post-CR-aware: $108k-$142.5k (64-84% of $170k SOW), up from $46k-$67k original target.")
        break

ws3.auto_filter.ref = None  # re-extend
ws3.auto_filter.ref = ws3.dimensions

# ===========================================================================
# Tab 4 — Borderline (update Commercial Approach + Resolution Buckets)
# ===========================================================================
ws4 = wb["4. Borderline"]

# Header row is 4 based on original: row 4 has "Ticket". Find it.
header_row = None
for r in range(1, 10):
    if str(ws4.cell(row=r, column=1).value or "").strip() == "Ticket":
        header_row = r
        break
header_row = header_row or 3

# Columns: A Ticket, B Summary, C Priority, D Lean, E Risk, F Commercial Approach
approach_updates = {
    "PAYM-151": "→ CR-3 (push-to-CR)",
    "PAYM-232": "→ CR-3 or Absorb",
    "PAYM-272": "→ CR-1 (permanent) — interim in scope",
    "PAYM-273": "→ CR-1 (midnight-crossing expansion)",
    "PAYM-293": "Split — bug absorb; CAG/PPL → CR-3",
    "PAYM-311": "→ CR-3",
    "PAYM-327": "Goodwill 50/50 (Simplification)",
    "PAYM-328": "Goodwill 50/50 (Simplification)",
    "PAYM-329": "COVERED BY CR02 (consolidation)",
    "PAYM-330": "Goodwill 50/50 (Simplification)",
    "PAYM-335": "Absorb (likely config defect)",
    "PAYM-337": "COVERED BY CR02 (email wrapper)",
    "PAYM-342": "COVERED BY CR02 + CR01 reversal",
    "PAYM-347": "Absorb (goodwill, trivial)",
    "PAYM-358": "COVERED BY CR01 (dual-cost-item template)",
}

for r in range(header_row + 1, ws4.max_row + 1):
    tk = ws4.cell(row=r, column=1).value
    if tk in approach_updates:
        cell = ws4.cell(row=r, column=6, value=approach_updates[tk])
        cell.alignment = WRAP
        cell.border = BORDER
        if "COVERED BY" in approach_updates[tk]:
            cell.fill = CR_COVERED
        elif "→ CR-4" in approach_updates[tk]:
            cell.fill = CR_NEW
        elif "Absorb" in approach_updates[tk] or "Goodwill" in approach_updates[tk]:
            cell.fill = BORDERLINE

# Update Resolution Buckets section
for r in range(1, ws4.max_row + 1):
    val = str(ws4.cell(row=r, column=1).value or "")
    if val == "Likely Absorb":
        ws4.cell(row=r, column=2, value="PAYM-272, 273, 335, 347 (+PAYM-329 now CR02)")
        ws4.cell(row=r, column=3, value=4)
        ws4.cell(row=r, column=5, value="CR02 pulls PAYM-329 into covered-by-CR. Remaining absorbs = low-cost defects/trivial.")
    elif val == "Negotiate Shared Cost":
        ws4.cell(row=r, column=2, value="PAYM-232, 293, 327, 328, 330 (+PAYM-358 now CR01, PAYM-342 now CR02)")
        ws4.cell(row=r, column=3, value=5)
        ws4.cell(row=r, column=5, value="PAYM-358 and PAYM-342 reclassified as covered by CR01/CR02 respectively. Simplification cluster is primary 50/50 negotiation.")
    elif val == "Push to CR":
        ws4.cell(row=r, column=2, value="PAYM-151, 311 (+PAYM-337 now CR02, moved)")
        ws4.cell(row=r, column=3, value=2)
        ws4.cell(row=r, column=5, value="PAYM-337 reclassified to CR02. PAYM-151 and PAYM-311 remaining push-to-CR.")

ws4.auto_filter.ref = ws4.dimensions

# ===========================================================================
# Tab 5 — Review Findings (add CR-aware pass + update strategy)
# ===========================================================================
ws5 = wb["5. Review Findings"]

# Title update
ws5["A1"] = "Finalised Review Findings and Commercial Strategy (CR-Aware Update)"
ws5["A2"] = "Independent multi-agent review — 16 April 2026; CR-aware update — 17 April 2026"

# Find methodology section and extend
# Methodology step 4 originally at row 8 → add step 5
method_insert_row = None
for r in range(1, ws5.max_row + 1):
    val = str(ws5.cell(row=r, column=1).value or "")
    if val == "4. Commercial Packaging":
        method_insert_row = r + 1
        break

if method_insert_row:
    ws5.insert_rows(method_insert_row, amount=2)
    ws5.cell(row=method_insert_row, column=1, value="5. CR Discovery (17/04/26)")
    ws5.cell(row=method_insert_row, column=2, value="Three new documents surfaced: CR01 Shift-Level Payroll Tax (raised 27/11/25, $16k); CR02 Additional Info in Pay Advices (raised 20/1/26, $6k); Project Evolution process doc (Dec 2025). 4-agent review pass.")
    ws5.cell(row=method_insert_row + 1, column=1, value="6. Commercial Reconciliation")
    ws5.cell(row=method_insert_row + 1, column=2, value="Revised recoverable from $46k-$67k to $108k-$142.5k (64-84% of SOW). Identified CR-4 Project Evolution Addendum ($31k-$50k) as net-new scope. Reclassified 5 tickets as 'Covered by CR'.")
    for r in (method_insert_row, method_insert_row + 1):
        for c in (1, 2):
            cell = ws5.cell(row=r, column=c)
            cell.alignment = WRAP
            cell.border = BORDER
            cell.fill = CR_NEW

# Add a new strategy section at the bottom
max_r5 = ws5.max_row
insert_at = max_r5 + 3
ws5.cell(row=insert_at, column=1, value="CR-AWARE NEGOTIATION STRATEGY (17/04/26)").font = SUB_FONT
insert_at += 1
ws5.cell(row=insert_at, column=1, value="#")
ws5.cell(row=insert_at, column=2, value="Recommendation")
for c in (1, 2):
    cell = ws5.cell(row=insert_at, column=c)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(wrap_text=True, vertical="center")
    cell.border = BORDER

new_strategies = [
    (1, "REWRITE CR02 before signing — remove 'more complex than initially identified from Discovery' phrasing. Discovery is 2c9 deliverable (SOW 3.2). Reframe as 'emergent requirement during build'."),
    (2, "COVERING LETTER on CR01 before countersignature — carve out: (i) month-end recalculation mode, (ii) PAYM-98 GL report build, (iii) SRO Part/All toggle."),
    (3, "Raise CR-4 Project Evolution Addendum BEFORE any Stage 4/5 development begins. Document Dec-2025 PE doc post-dates Nov-2025 CR01 as 'emergent post-CR'."),
    (4, "Maintain existing negotiation stack for CR-1/CR-2/CR-3 — CR01/CR02 do NOT reduce those positions. CR01 is net-new; CR02 covers 5 tickets but the other 14 scope creep + 10 borderline remain."),
    (5, "Timing: CR02-rewrite + CR01-covering-letter THIS WEEK → CR-1/CR-2 (2 wks, before PPT2) → CR-4 (2-4 wks) → CR-3 (post-PPT2)."),
    (6, "Precedent risk: if Hireup accept CR02's 'Discovery under-scoped' framing, it becomes reusable for CR-1/CR-3. If they reject it, damages CR-1/CR-3. REMOVE the phrase to protect the stack."),
    (7, "Cite SOW §5 Assumption 8 in CR-4 covering note — same logic that excludes Hibiscus supports excluding PE Stages 4-5."),
    (8, "Revised target recoverable: $108k-$142.5k (64-84% of $170k SOW). Depend on (a) CR02 rewrite, (b) CR01 carve-out letter, (c) CR-4 raised pre-dev."),
]
for i, (num, rec) in enumerate(new_strategies, start=insert_at + 1):
    ws5.cell(row=i, column=1, value=num)
    ws5.cell(row=i, column=2, value=rec)
    for c in (1, 2):
        cell = ws5.cell(row=i, column=c)
        cell.alignment = WRAP
        cell.border = BORDER

# Widen column B on this tab
ws5.column_dimensions["B"].width = 110

# Save
wb.save(DST)
print(f"WROTE {DST}")
