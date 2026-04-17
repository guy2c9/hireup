"""Build the CR-aware reconciliation xlsx — 5 tabs, formatted."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = "/Users/guybaxter/Documents/Claude Code/hireup/docs/Hireup SCHADS — CR-Aware Commercial Reconciliation.xlsx"

HEADER_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Calibri", size=16, bold=True, color="1F3864")
SUB_FONT = Font(name="Calibri", size=11, bold=True, color="1F3864")
BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)
WRAP = Alignment(wrap_text=True, vertical="top")

CR_RAISED = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
CR_NEW = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
CR_PRIORITY = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
RISK_HIGH = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
RISK_MED = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
RISK_LOW = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
SCOPE_CREEP = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
BORDERLINE = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
IN_SCOPE = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")


def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")
        cell.border = BORDER


def auto_width(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def apply_rows(ws, data, start_row, fill_map=None):
    for r, row in enumerate(data, start=start_row):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = WRAP
            cell.border = BORDER
            cell.font = Font(name="Calibri", size=10)
            if fill_map:
                fill = fill_map(row, c)
                if fill:
                    cell.fill = fill


wb = openpyxl.Workbook()
wb.remove(wb.active)

# ===== Tab 1: Executive Summary =====
ws = wb.create_sheet("1. Executive Summary")
ws.merge_cells("A1:F1")
ws["A1"] = "Hireup SCHADS — CR-Aware Commercial Reconciliation"
ws["A1"].font = TITLE_FONT
ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[1].height = 28

meta = [
    ("Project", "Hireup Payroll Implementation"),
    ("SOW Reference", "SOW No. 20250717H"),
    ("SOW Value", "$170,000 (fixed price)"),
    ("Analysis Date", "17 April 2026"),
    ("Prepared By", "2cloudnine"),
    ("Review Method", "4-agent independent analysis (CR01, CR02, Project Evolution, Commercial Reconciliation)"),
    ("Supplements", "Hireup Scope Analysis — Finalised.xlsx (97-ticket baseline)"),
]
for i, (k, v) in enumerate(meta, start=3):
    ws.cell(row=i, column=1, value=k).font = SUB_FONT
    ws.cell(row=i, column=2, value=v)

ws.cell(row=11, column=1, value="REVISED COMMERCIAL POSITION").font = SUB_FONT
headers = ["CR", "Title", "Status", "Tickets", "$ Low", "$ High", "Priority", "Contractual Basis"]
for c, h in enumerate(headers, 1):
    ws.cell(row=12, column=c, value=h)
style_header(ws, 12, len(headers))

cr_rows = [
    ("CR01", "Shift-Level Payroll Tax & FP/NFP Cost Item", "Raised 27/11/25 — awaiting signature", "Adjacent: 98, 115, 145, 342, 357, 358", 16000, 16000, "P0 — sign with carve-out letter", "Net-new scope (no ticket coverage). Flat fee. True effort $24k–$30k."),
    ("CR02", "Additional Information in Pay Advices", "Raised 20/1/26 — REWRITE before signing", "PAYM-167, 329, 337, 342, 123", 6000, 6000, "P0 — rewrite urgently", "Under-priced ($6k vs $15k–$22.5k realistic). 'Discovery under-scoped' language is contractual risk."),
    ("CR02-Uplift", "CR02 rewrite to realistic T&M", "Recommended (not raised)", "Same as CR02", 9000, 16500, "P1 — bundle with CR02 rewrite", "Closes gap to 60–90 hrs. Removes Discovery-fault narrative."),
    ("CR-1", "Sleepover Engine Extension", "Not raised", "PAYM-45, 272, 273, 312, 313, 319, 324, 325, 326", 20000, 25000, "P0 — raise before PPT2", "SOW Gap 2 narrow: 'compliance hours at lower rate'. Min top-ups, 18+hr, back-to-back excluded."),
    ("CR-2", "Portable LSL Compliance", "Not raised", "PAYM-159, 173, 241", 25000, 35000, "P0 — raise before PPT2", "Zero ambiguity — not in SOW, RW, or DW. Strongest defensibility."),
    ("CR-3", "Enhancements & Integrations Bundle", "Not raised", "PAYM-166, 262, 271, 279, 280, 333, 334, 354 (+ 151/311)", 14000, 20000, "P1 — raise after CR-1/CR-2", "PAYM-334 already built — reconcile at 50–60% rate."),
    ("CR-4", "Project Evolution — Month-End & Audit Trail (NEW)", "Not raised — RECOMMENDED", "Stages 4–5: month-end recalc, SRO toggle, correction/linkage", 31000, 50000, "P0 — raise BEFORE Stage 4/5 dev", "Dec-2025 design post-dates CR01 (Nov-2025). Emergent post-CR scope."),
    ("SUB-TOTAL", "Potential recoverable (gross)", "", "", 121000, 162500, "", ""),
    ("Concessions", "Goodwill reapplied (50/50 Simplification + absorb items)", "", "PAYM-327/328/330 50/50; 335, 347 absorb", -13000, -20000, "", "Relationship investment"),
    ("NET", "Net Recoverable Total", "", "", 108000, 142500, "", "64–84% of $170k SOW"),
]

for r, row in enumerate(cr_rows, start=13):
    for c, val in enumerate(row, start=1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.alignment = WRAP
        cell.border = BORDER
        cell.font = Font(name="Calibri", size=10)
        if c in (5, 6) and isinstance(val, (int, float)):
            cell.number_format = '"$"#,##0'
        status = row[2] if len(row) > 2 else ""
        if "RECOMMENDED" in str(status) or "NEW" in str(row[0]) or row[0] == "CR-4":
            cell.fill = CR_NEW
        elif "Raised" in str(status):
            cell.fill = CR_RAISED
        elif row[0] in ("SUB-TOTAL", "NET", "Concessions"):
            cell.fill = HEADER_FILL
            cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        elif "P0" in str(row[6] if len(row) > 6 else ""):
            cell.fill = CR_PRIORITY

ws.cell(row=25, column=1, value="HEADLINE").font = SUB_FONT
ws.merge_cells("A26:H29")
ws["A26"] = (
    "Original strategy targeted $46k–$67k recoverable. Post-CR-aware analysis: $108k–$142.5k net recoverable "
    "(64–84% of $170k SOW) — contingent on (a) rewriting CR02 this week to remove 'Discovery under-scoped' "
    "language, (b) covering-letter carve-out on CR01 before countersignature, (c) raising CR-4 before any "
    "Stage 4/5 development begins. The two most load-bearing actions are CR02 rewrite and CR01 covering "
    "letter — both this week, before PPT2 gates."
)
ws["A26"].alignment = WRAP
ws["A26"].font = Font(name="Calibri", size=11, italic=True)
auto_width(ws, [14, 38, 30, 38, 12, 12, 32, 50])

# ===== Tab 2: CR Register =====
ws2 = wb.create_sheet("2. CR Register")
ws2.append([
    "CR ID", "Title", "Status", "Date Raised", "Date Signed", "$ Low", "$ High",
    "Agreed $", "Ticket Count", "Priority", "Contractual Basis", "Notes"
])
style_header(ws2, 1, 12)

cr_register = [
    ["CR01", "Shift-Level Payroll Tax & FP/NFP Cost Item", "Awaiting signature", "27/11/2025", "",
     16000, 16000, "", 0, "P0",
     "Net-new post-SOW scope. No JIRA ticket coverage in 97-ticket set.",
     "Flat $10k dev + $6k T&M (24 hrs). Under-priced; true effort $24k–$30k. ISSUE COVERING LETTER before countersignature carving out month-end recalc and PAYM-98 GL report."],
    ["CR02", "Additional Information in Pay Advices", "Awaiting signature — REWRITE FIRST", "20/01/2026", "",
     6000, 6000, "", 5, "P0",
     "Gap 7 rescope + net-new shift-level data sourcing + layout redesign.",
     "$6k for 24 hrs T&M. True effort 60–100 hrs ($15k–$22.5k). 'Significantly more complex than initially identified from Discovery' phrasing is a contractual gift to Hireup — Discovery is 2c9 deliverable (SOW 3.2). REWRITE required."],
    ["CR02-Uplift", "CR02 rewrite uplift to realistic T&M", "Recommended — not raised", "", "",
     9000, 16500, "", 0, "P1",
     "Bundled with CR02 rewrite.",
     "Closes gap between raised $6k and realistic $15k–$22.5k. Reframes rationale as 'emergent requirement during build' rather than Discovery failure."],
    ["CR-1", "Sleepover Engine Extension", "Not raised", "", "",
     20000, 25000, "", 9, "P0",
     "SOW Gap 2 ('compliance hours at the lower rate') narrowly worded — excludes min top-ups, 18+hr, back-to-back, midday penalties, OT aggregation.",
     "Raise BEFORE PPT2 completes. Strongest defensibility after CR-2."],
    ["CR-2", "Portable LSL Compliance", "Not raised", "", "",
     25000, 35000, "", 3, "P0",
     "Zero ambiguity — not in SOW, RW, DW, or RAID.",
     "Highest-value CR; zero-risk contractual position. Raise BEFORE PPT2 completes."],
    ["CR-3", "Enhancements & Integrations Bundle", "Not raised", "", "",
     14000, 20000, "", 10, "P1",
     "Mixed scope creep + borderline-CR items.",
     "Negotiation-flex CR. Include PAYM-334 (already built) at 50–60% rate. Offer 10–15% package discount if all CRs approved together."],
    ["CR-4", "Project Evolution — Month-End & Audit Trail", "NEW — recommended, not raised", "", "",
     31000, 50000, "", 0, "P0",
     "Dec-2025 Project Evolution doc introduces Stages 4–5 (month-end recalc engine, SRO Part/All toggle, reversal→correction→linkage audit trail) beyond CR01.",
     "Timing: post-Nov 2025 CR01. 'Emergent post-CR'. Cite SOW §5 Assumption 8 parallel. Raise BEFORE any Stage 4/5 dev starts."],
]

for r, row in enumerate(cr_register, start=2):
    for c, val in enumerate(row, start=1):
        cell = ws2.cell(row=r, column=c, value=val)
        cell.alignment = WRAP
        cell.border = BORDER
        cell.font = Font(name="Calibri", size=10)
        if c in (6, 7):
            cell.number_format = '"$"#,##0'
        if "NEW" in str(row[2]):
            cell.fill = CR_NEW
        elif "Awaiting" in str(row[2]):
            cell.fill = CR_RAISED
        elif row[9] == "P0":
            cell.fill = CR_PRIORITY

auto_width(ws2, [12, 40, 30, 13, 13, 11, 11, 11, 10, 10, 40, 60])
ws2.freeze_panes = "A2"
ws2.auto_filter.ref = ws2.dimensions

# ===== Tab 3: Ticket → CR Map =====
ws3 = wb.create_sheet("3. Ticket–CR Map")
ws3.append([
    "Ticket", "Summary", "Original Class", "Revised Class", "Assigned CR",
    "Coverage", "Confidence", "Rationale", "Carve-out Flag"
])
style_header(ws3, 1, 9)

ticket_map = [
    # Scope creep (19)
    ["PAYM-45", "Minimum Sleepover not working as expected", "Scope Creep", "Scope Creep", "CR-1", "Direct", "H", "Gap 2 narrow wording — min top-up bespoke to IFA", ""],
    ["PAYM-98", "Finance GL Report Build — Location requirements", "Scope Creep", "Scope Creep (data model via CR01)", "CR-4 / CR-3", "Partial", "H", "CR01 supplies shift-Location data; report build still unfunded", "CARVE OUT from CR01"],
    ["PAYM-159", "Portable LSL custom field requirements", "Scope Creep", "Scope Creep", "CR-2", "Direct", "H", "Not in SOW/RW/DW", ""],
    ["PAYM-166", "Slack integration", "Scope Creep", "Scope Creep", "CR-3", "Direct", "H", "Tagged 'Nice to have' — outside scope", ""],
    ["PAYM-173", "Complete Portable LSL Workbook", "Scope Creep", "Scope Creep", "CR-2", "Direct", "H", "PLSL compliance stream", ""],
    ["PAYM-241", "Portable LSL Reports (5 states)", "Scope Creep", "Scope Creep", "CR-2", "Direct", "H", "PLSL-specific reporting; not covered by CR02 pay advice work", ""],
    ["PAYM-262", "Hibiscus Release adoption", "Scope Creep", "Scope Creep", "CR-3", "Direct", "H", "SOW §5 Assumption 8 excludes future releases", ""],
    ["PAYM-271", "Minimum Contracted Hours (PPT)", "Scope Creep", "Scope Creep", "CR-3", "Direct", "H", "JIRA comments pre-acknowledge CR", ""],
    ["PAYM-279", "FDVL & Pay advices suppression", "Scope Creep", "Scope Creep", "CR-3", "Direct", "H", "Compliance rule distinct from CR02 display redesign", ""],
    ["PAYM-280", "New pay code process tooling", "Scope Creep", "Scope Creep", "CR-3", "Direct", "H", "Ops tooling outside SOW", ""],
    ["PAYM-312", "Sleepover >18hr", "Scope Creep", "Scope Creep", "CR-1", "Direct", "H", "Gap 2 narrow", ""],
    ["PAYM-313", "Sleepover >18hr back-to-back", "Scope Creep", "Scope Creep", "CR-1", "Direct", "H", "Gap 2 narrow", ""],
    ["PAYM-319", "Minimum Sleepover Rates", "Scope Creep", "Scope Creep", "CR-1", "Direct", "H", "Gap 2 narrow", ""],
    ["PAYM-324", "Daily OT Day + Sleepover", "Scope Creep", "Scope Creep", "CR-1", "Direct", "H", "OT aggregation not in Gap 2", ""],
    ["PAYM-325", "Sleepover crossing midday", "Scope Creep", "Scope Creep", "CR-1", "Direct", "H", "Midday penalty not in Gap 2", ""],
    ["PAYM-326", "New pay code Min Sleepover", "Scope Creep", "Scope Creep", "CR-1", "Direct", "H", "Min top-up pay code bespoke", ""],
    ["PAYM-333", "Apex/flow audit", "Scope Creep", "Scope Creep", "CR-3", "Direct", "M", "Ops monitoring outside SOW", ""],
    ["PAYM-334", "Custom UI Widget (already built)", "Scope Creep", "Scope Creep — reduced rate", "CR-3 @ 50–60%", "Direct", "H", "Already consumed effort; first concession", ""],
    ["PAYM-354", "Production API Limits monitoring", "Scope Creep", "Scope Creep", "CR-3", "Direct", "M", "Ops monitoring", ""],
    # Borderline (15)
    ["PAYM-151", "Friday Booking reconciliation report", "Borderline", "Borderline → CR", "CR-3", "Direct", "M", "Custom report, not covered by CR01/02", ""],
    ["PAYM-232", "Auto-approve workflow", "Borderline", "Borderline → Negotiate/Absorb", "CR-3 or Absorb", "Partial", "L", "Custom scheduled automation; partial scope basis", ""],
    ["PAYM-272", "Sleepover midday workaround", "Borderline", "Borderline → CR (permanent)", "CR-1", "Partial", "M", "Interim in scope; permanent fix is CR-1 territory", ""],
    ["PAYM-273", "Sleepover first & active", "Borderline", "Borderline → CR", "CR-1", "Partial", "M", "Midnight-crossing expansion beyond Gap 1", ""],
    ["PAYM-293", "Pay trans $34.59 defect", "Borderline", "Split — bug absorb; rates CR", "Absorb / CR-3", "Partial", "M", "Defect portion absorb; CAG/PPL rates CR-3", ""],
    ["PAYM-311", "CAG/PPL default rates", "Borderline", "Borderline → CR", "CR-3", "Direct", "M", "Not in DW", ""],
    ["PAYM-327", "Simplification 1 — Pay Code", "Borderline", "Goodwill 50/50", "Shared", "Direct", "H", "SOW commits one build cycle; shared accountability", ""],
    ["PAYM-328", "Simplification 2 — Sleepover/Remote", "Borderline", "Goodwill 50/50", "Shared", "Direct", "H", "Shared accountability", ""],
    ["PAYM-329", "Simplification 3 — Discretionary Allowance", "Borderline → Absorb", "Covered by CR02", "CR02", "Direct", "H", "Agent B: part of 4-component consolidation", ""],
    ["PAYM-330", "Simplification 4 — Broken/Split", "Borderline", "Goodwill 50/50", "Shared", "Direct", "H", "Shared accountability", ""],
    ["PAYM-335", "Leave events adjustments", "Borderline", "Absorb (likely config defect)", "Absorb", "Direct", "M", "Defect territory", ""],
    ["PAYM-337", "Custom Payslip Email body", "Borderline → CR", "Covered by CR02", "CR02", "Partial", "M", "Email wrapper likely bundled with layout redesign", ""],
    ["PAYM-342", "Reversed Transactions on Pay Advice", "Borderline → Negotiate", "Covered by CR02 (primary)", "CR02 + CR01 adjacent", "Direct", "M", "CR02 layout; CR01 reversal mechanism", ""],
    ["PAYM-347", "MV Allowance $1.04/km", "Borderline → Absorb", "Goodwill Absorb", "Absorb", "Direct", "H", "Trivial precedent; absorb", ""],
    # CR01-adjacent reclassifications
    ["PAYM-115", "[Test] GL Integration (Netsuite)", "In Scope", "In Scope — complexity uplift", "CR01 adjacent", "Adjacent", "M", "FP/NFP + reversal journals grow scope", "CARVE OUT complexity in CR-4"],
    ["PAYM-123", "SACS Project Codes", "In Scope", "In Scope — CR02 adjacent", "CR02 adjacent", "Adjacent", "L", "Touches pay-code consolidation", ""],
    ["PAYM-145", "Change of State and LSL", "In Scope", "In Scope — split CR01/CR-2", "CR01 + CR-2", "Partial", "M", "State changes first-class via CR01; LSL accrual/PLSL linkage to CR-2", ""],
    ["PAYM-167", "Hireup custom Payslip template", "In Scope (Gap 7)", "Covered by CR02", "CR02", "Direct", "H", "Agent B: CR02 core target", ""],
    ["PAYM-357", "Review and check GL Account codes", "In Scope", "In Scope — CR01 uplift", "CR01 adjacent", "Adjacent", "M", "FP/NFP doubles GL surface", ""],
    ["PAYM-358", "Training GL split (50710/50720)", "Borderline → Negotiate", "Covered by CR01 (template)", "CR01 adjacent", "Partial", "M", "Dual-cost-item pattern covers", ""],
]

def ticket_fill(row, c):
    rc = str(row[3])
    if c == 4:
        if "Scope Creep" in rc:
            return SCOPE_CREEP
        if "Borderline" in rc or "Goodwill" in rc or "Absorb" in rc:
            return BORDERLINE
        if "In Scope" in rc or "Covered" in rc:
            return IN_SCOPE
    if c == 9 and "CARVE OUT" in str(row[8]):
        return RISK_HIGH
    return None

apply_rows(ws3, ticket_map, 2, ticket_fill)
auto_width(ws3, [11, 40, 22, 32, 20, 12, 10, 50, 22])
ws3.freeze_panes = "A2"
ws3.auto_filter.ref = ws3.dimensions

# ===== Tab 4: Commercial Actions =====
ws4 = wb.create_sheet("4. Commercial Actions")
ws4.append([
    "#", "Action", "Owner", "Timing", "Status", "$ Impact Low", "$ Impact High",
    "Dependencies", "Notes"
])
style_header(ws4, 1, 9)

actions = [
    [1, "Withdraw CR02 draft; rewrite at 60–90 hrs ($15k–$22.5k); strip 'Discovery under-scoped' phrasing",
     "Guy / 2c9 Delivery Lead", "This week — before Hireup signs", "Open", 9000, 16500, "None", "Reframe as 'emergent requirement during build'"],
    [2, "Sign CR01 but issue covering letter carving out (i) month-end recalc, (ii) PAYM-98 GL report build, (iii) SRO Part/All toggle",
     "2c9 Delivery Lead", "Before CR01 countersignature", "Open", 31000, 50000, "CR01 not yet signed", "Protects CR-4 position"],
    [3, "Raise CR-4 Project Evolution Addendum (Stages 4–5)",
     "Guy + Delivery Lead", "Within 2 weeks, before any Stage 4/5 dev", "Open", 31000, 50000, "Action #2 covering letter", "Cite Dec-2025 doc post-dates Nov-2025 CR01"],
    [4, "Raise CR-1 Sleepover Engine ($20k–$25k)",
     "Account Lead", "Within 2 weeks — before PPT2", "Open", 20000, 25000, "None", "SOW Gap 2 narrow citation"],
    [5, "Raise CR-2 Portable LSL ($25k–$35k)",
     "Account Lead", "Within 2 weeks — before PPT2", "Open", 25000, 35000, "None", "Zero ambiguity — strongest position"],
    [6, "Raise CR-3 Enhancements Bundle; PAYM-334 at 50–60% rate",
     "Account Lead", "After CR-1/CR-2 lodged", "Open", 14000, 20000, "Actions #4, #5", "Package discount 10–15% if all approved together"],
    [7, "Document in writing that CR01's 24hr T&M did NOT cover correction/linkage or month-end recalc",
     "2c9 Delivery Lead", "This week", "Open", 0, 0, "None", "Defends CR-4 commercial position"],
    [8, "Apply goodwill concessions: 50/50 on PAYM-327/328/330; absorb PAYM-347, 335",
     "Account Lead", "With CR-1/CR-2 lodgement", "Open", -13000, -20000, "Actions #4, #5", "Relationship capital to strengthen CR positions"],
    [9, "Reclassify in Finalised sheet: 167, 329, 337, 342 → 'Covered by CR02'; 358 → 'Covered by CR01'",
     "Guy", "Within 1 week", "Open", 0, 0, "None", "Housekeeping — keep Finalised in sync"],
    [10, "Sequence: CR02-rewrite + CR01-letter (now) → CR-1/CR-2 (2 wks) → CR-4 (2–4 wks) → CR-3 (post-PPT2)",
     "Guy + Account Lead", "Ongoing", "In Progress", 0, 0, "All", "Preserves negotiation leverage"],
]

def action_fill(row, c):
    if c == 5 and str(row[4]) in ("Open",):
        return RISK_MED
    if str(row[3]).startswith("This week"):
        return CR_PRIORITY if c == 4 else None
    return None

apply_rows(ws4, actions, 2, action_fill)
for r in range(2, 2 + len(actions)):
    for c in (6, 7):
        cell = ws4.cell(row=r, column=c)
        if isinstance(cell.value, (int, float)):
            cell.number_format = '"$"#,##0'
auto_width(ws4, [4, 60, 22, 28, 12, 12, 12, 22, 44])
ws4.freeze_panes = "A2"
ws4.auto_filter.ref = ws4.dimensions

# ===== Tab 5: Risk Register =====
ws5 = wb.create_sheet("5. Risk Register")
ws5.append([
    "#", "Risk", "Severity", "Linked CR", "Mitigation", "Owner",
    "Status", "Date Identified"
])
style_header(ws5, 1, 8)

risks = [
    [1, "CR02 'more complex than Discovery identified' language — Discovery is 2c9 deliverable (SOW 3.2). Hireup counsel could argue under-estimation is 2c9 risk to absorb.",
     "HIGH", "CR02", "Withdraw, rewrite as 'emergent requirement during build'. Never admit Discovery fault in writing.",
     "Guy / Delivery Lead", "Open", "17/04/2026"],
    [2, "CR02 precedent risk — if Hireup accepts 'Discovery under-scoped' framing, reusable for CR-1/CR-3. If rejected, damages CR-1/CR-3.",
     "HIGH", "CR02 → CR-1, CR-3", "Remove phrase entirely. Reserve rescope argument in covering note.",
     "Guy", "Open", "17/04/2026"],
    [3, "CR01 'journal reversal capability' could be read broadly to subsume Stage 5 correction/linkage and Stage 4 recalc.",
     "HIGH", "CR01 / CR-4", "Covering letter NOW: explicit written carve-out of month-end recalc, correction journals, 3-way audit linkage.",
     "Delivery Lead", "Open", "17/04/2026"],
    [4, "CR01 $16k flat materially under-priced ($24k–$30k true) — sets low-ball precedent.",
     "MED", "CR01", "Frame CR-4 as 'uncontemplated by CR01 pricing'; use Agent A effort estimate as benchmark.",
     "Delivery Lead", "Open", "17/04/2026"],
    [5, "Stages 4/5 drift — Dec-2025 PE doc adds scope Hireup may assume is 'included' since CR01 touched PE.",
     "HIGH", "CR01 / CR-4", "CR-4 raised BEFORE any dev. Document Dec-2025 post-dates Nov-2025 CR01.",
     "Guy + Delivery Lead", "Open", "17/04/2026"],
    [6, "PAYM-334 already built — no commercial mechanism to recover.",
     "MED", "CR-3", "Include in CR-3 at 50–60% reduced rate. First concession if negotiation needed.",
     "Account Lead", "Open", "17/04/2026"],
    [7, "SOW §5 Assumption 8 parallel — excludes Hibiscus AND supports excluding Stages 4/5 (positive risk).",
     "LOW", "CR-3 / CR-4", "Cite explicitly in CR-4 and CR-3 cover notes.",
     "Guy", "Mitigated", "17/04/2026"],
    [8, "Gap 7 own-mouth admission via CR02 — '2c9 said on paper' that Gap 7 was under-specified.",
     "MED", "CR02", "Reframe as 'evolved scope emerged during build'; do NOT characterise as Discovery failure.",
     "Guy", "Open", "17/04/2026"],
    [9, "Simplification cluster (PAYM-327–330) — rework of initial designs; SOW commits to one build cycle.",
     "MED", "CR-3 / Shared", "50/50 share as goodwill; avoid precedent of full absorption.",
     "Account Lead", "Open", "17/04/2026"],
    [10, "PPT2 timing pressure — raising CRs after PPT2 completes weakens position.",
     "HIGH", "All CRs", "All CRs lodged before PPT2 gate. Calendar the deadline now.",
     "Guy + Account Lead", "Open", "17/04/2026"],
]

def risk_fill(row, c):
    if c == 3:
        s = str(row[2])
        if "HIGH" in s:
            return RISK_HIGH
        if "MED" in s:
            return RISK_MED
        if "LOW" in s:
            return RISK_LOW
    return None

apply_rows(ws5, risks, 2, risk_fill)
auto_width(ws5, [4, 70, 11, 22, 62, 20, 12, 15])
ws5.freeze_panes = "A2"
ws5.auto_filter.ref = ws5.dimensions

wb.save(OUT)
print(f"WROTE {OUT}")
