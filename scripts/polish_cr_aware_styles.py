"""Polish the CR-Aware Update xlsx — normalise fonts, colours, and spacing.

Matches the original Finalised.xlsx palette:
  - Titles:            size 14, bold, no fill
  - Subtitles:         size 10, regular
  - Section banners:   size 12, bold, fill D9E2F3 (light blue), dark text
  - Table headers:     size 11, bold, white text, fill 073763 (dark navy)
  - Body cells:        size 10, regular, wrap, top-aligned
  - Fills:
      In Scope  → D9EAD3 (light green)
      Creep     → F4CCCC (light red)
      Border    → FCE5CD (light orange)
      CR-raised → FFF2CC (yellow — distinct, for CR01/CR02 raised rows)
      CR-new    → FCE4D6 (peach — distinct, for CR-4 / Uplift / recommended)
      CR-cover  → BDD7EE (light blue — ticket covered by existing CR)
"""
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from copy import copy

DST = Path("/Users/guybaxter/Documents/Claude Code/hireup/docs/Hireup Scope Analysis — Finalised (CR-Aware Update).xlsx")

TITLE = Font(name="Calibri", size=14, bold=True, color="000000")
SUBTITLE = Font(name="Calibri", size=10, color="000000")
SECTION = Font(name="Calibri", size=12, bold=True, color="073763")
TABLE_HDR = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
BODY = Font(name="Calibri", size=10, color="000000")
BODY_BOLD = Font(name="Calibri", size=10, bold=True, color="000000")
META_LABEL = Font(name="Calibri", size=11, bold=True, color="000000")

FILL_SECTION = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
FILL_HEADER = PatternFill(start_color="073763", end_color="073763", fill_type="solid")
FILL_IN_SCOPE = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
FILL_CREEP = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
FILL_BORDER = PatternFill(start_color="FCE5CD", end_color="FCE5CD", fill_type="solid")
FILL_CR_RAISED = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
FILL_CR_NEW = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
FILL_CR_COVERED = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
FILL_NONE = PatternFill(fill_type=None)

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP_TOP = Alignment(wrap_text=True, vertical="top")
WRAP_CENTER = Alignment(wrap_text=True, vertical="center")


def style_body(cell):
    cell.font = BODY
    cell.alignment = WRAP_TOP
    cell.border = BORDER


def style_section_header(cell):
    cell.font = SECTION
    cell.fill = FILL_SECTION
    cell.alignment = WRAP_CENTER
    cell.border = BORDER


def style_table_header(cell):
    cell.font = TABLE_HDR
    cell.fill = FILL_HEADER
    cell.alignment = WRAP_CENTER
    cell.border = BORDER


def classify_fill_from_text(text):
    t = str(text or "")
    if "Covered by CR" in t or "COVERED BY" in t:
        return FILL_CR_COVERED
    if "Scope Creep" in t:
        return FILL_CREEP
    if "Borderline" in t:
        return FILL_BORDER
    if "In Scope" in t:
        return FILL_IN_SCOPE
    return None


wb = openpyxl.load_workbook(DST)

# ============================================================================
# Tab 1 — Executive Summary
# ============================================================================
ws = wb["1. Executive Summary"]

# Row 1 — title
ws["A1"].font = TITLE
ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[1].height = 24

# Meta block: rows 3-9 (Project, SOW Ref, SOW Value, SOW Date, Analysis Date, Prepared By, Review Method)
for r in range(3, 10):
    label = ws.cell(row=r, column=1)
    value = ws.cell(row=r, column=2)
    if label.value:
        label.font = META_LABEL
        label.alignment = WRAP_TOP
    if value.value:
        value.font = BODY
        value.alignment = WRAP_TOP

# Classification Summary block: row 10 banner, row 11 headers, rows 12-15 data
if str(ws.cell(row=10, column=1).value or "") == "CLASSIFICATION SUMMARY":
    style_section_header(ws.cell(row=10, column=1))
    ws.merge_cells(start_row=10, start_column=1, end_row=10, end_column=4)
    for c in range(1, 4):
        style_table_header(ws.cell(row=11, column=c))

# Commercial Exposure Summary — row 16 banner
if str(ws.cell(row=16, column=1).value or "") == "COMMERCIAL EXPOSURE SUMMARY":
    style_section_header(ws.cell(row=16, column=1))
    ws.merge_cells(start_row=16, start_column=1, end_row=16, end_column=4)
# row 17 headers
for c in range(1, 5):
    cell = ws.cell(row=17, column=c)
    if cell.value:
        style_table_header(cell)
# rows 18-26 — CR exposure rows (styled by builder; re-apply consistent body font)
for r in range(18, 27):
    label = str(ws.cell(row=r, column=1).value or "")
    for c in range(1, 5):
        cell = ws.cell(row=r, column=c)
        if cell.value is None:
            continue
        cell.font = BODY
        cell.alignment = WRAP_TOP
        cell.border = BORDER
        if isinstance(cell.value, (int, float)):
            cell.number_format = '"$"#,##0;[Red]-"$"#,##0'
    # Re-fill rows based on status
    if "raised 27/11" in label or "raised 20/1" in label:
        fill = FILL_CR_RAISED
    elif "NEW RECOMMENDED" in label or "Uplift" in label:
        fill = FILL_CR_NEW
    elif label == "Net Recoverable Total":
        fill = FILL_HEADER
    else:
        fill = None
    if fill:
        for c in range(1, 5):
            cell = ws.cell(row=r, column=c)
            if cell.value is not None:
                cell.fill = fill
                if fill is FILL_HEADER:
                    cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")

# Recommended CR Packaging — find banner row 28
if "RECOMMENDED CR PACKAGING" in str(ws.cell(row=28, column=1).value or ""):
    style_section_header(ws.cell(row=28, column=1))
    ws.merge_cells(start_row=28, start_column=1, end_row=28, end_column=4)
# row 29 headers
for c in range(1, 5):
    style_table_header(ws.cell(row=29, column=c))
# rows 30-37 — CR rows
for r in range(30, 38):
    cr_id = str(ws.cell(row=r, column=1).value or "")
    for c in range(1, 5):
        cell = ws.cell(row=r, column=c)
        if cell.value is None:
            continue
        cell.font = BODY
        cell.alignment = WRAP_TOP
        cell.border = BORDER
    if cr_id in ("CR01", "CR02"):
        fill = FILL_CR_RAISED
    elif cr_id in ("CR02-Uplift", "CR-4"):
        fill = FILL_CR_NEW
    else:
        fill = None
    if fill:
        for c in range(1, 5):
            if ws.cell(row=r, column=c).value is not None:
                ws.cell(row=r, column=c).fill = fill

# Key Findings — banner row 39
if "KEY FINDINGS" in str(ws.cell(row=39, column=1).value or ""):
    style_section_header(ws.cell(row=39, column=1))
    ws.merge_cells(start_row=39, start_column=1, end_row=39, end_column=4)
# row 40 headers
for c in range(1, 4):
    style_table_header(ws.cell(row=40, column=c))
# rows 41-50 — findings
for r in range(41, 51):
    for c in range(1, 4):
        cell = ws.cell(row=r, column=c)
        if cell.value is None:
            continue
        cell.font = BODY
        cell.alignment = WRAP_TOP
        cell.border = BORDER

# Top Actions — banner
for r in range(51, ws.max_row + 1):
    val = str(ws.cell(row=r, column=1).value or "")
    if "TOP ACTIONS" in val:
        style_section_header(ws.cell(row=r, column=1))
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        # headers next row
        for c in range(1, 4):
            style_table_header(ws.cell(row=r + 1, column=c))
        # data rows
        for rr in range(r + 2, r + 12):
            for c in range(1, 4):
                cell = ws.cell(row=rr, column=c)
                if cell.value is None:
                    continue
                cell.font = BODY
                cell.alignment = WRAP_TOP
                cell.border = BORDER
                timing = str(ws.cell(row=rr, column=3).value or "")
                if "This week" in timing:
                    cell.fill = FILL_CR_NEW
        break

# Column widths (match original A=30 but extend to fit new content)
ws.column_dimensions["A"].width = 56
ws.column_dimensions["B"].width = 65
ws.column_dimensions["C"].width = 22
ws.column_dimensions["D"].width = 32

# Uniform row heights for long text rows
for r in range(18, ws.max_row + 1):
    has_wrap = any(
        ws.cell(row=r, column=c).value and len(str(ws.cell(row=r, column=c).value)) > 80
        for c in range(1, 5)
    )
    if has_wrap:
        ws.row_dimensions[r].height = 42

# ============================================================================
# Tab 2 — All Tickets
# ============================================================================
ws2 = wb["2. All Tickets"]

# Header row
for c in range(1, ws2.max_column + 1):
    cell = ws2.cell(row=1, column=c)
    if cell.value:
        style_table_header(cell)
ws2.row_dimensions[1].height = 32

# Body rows — restyle everything consistently
for r in range(2, ws2.max_row + 1):
    classification = str(ws2.cell(row=r, column=6).value or "")
    for c in range(1, ws2.max_column + 1):
        cell = ws2.cell(row=r, column=c)
        if cell.value is None and c != 7:
            continue
        cell.font = BODY
        cell.alignment = WRAP_TOP
        cell.border = BORDER

    # Reapply row fill based on classification (columns A-F get class fill;
    # column G gets its own fill based on coverage)
    row_fill = classify_fill_from_text(classification)
    if row_fill:
        for c in range(1, 7):
            ws2.cell(row=r, column=c).fill = row_fill

    # Column G — coverage status
    coverage = str(ws2.cell(row=r, column=7).value or "")
    cov_fill = None
    if "COVERED BY" in coverage:
        cov_fill = FILL_CR_COVERED
    elif "→ CR-4" in coverage or "NEW" in coverage:
        cov_fill = FILL_CR_NEW
    elif "Goodwill" in coverage or "Absorb" in coverage:
        cov_fill = FILL_BORDER
    elif coverage.startswith("→"):
        cov_fill = FILL_CR_RAISED if "CR01" in coverage or "CR02" in coverage else None
    if cov_fill:
        ws2.cell(row=r, column=7).fill = cov_fill

    # Columns H (Source References) and I (Explanation) keep white background
    for c in (8, 9):
        cell = ws2.cell(row=r, column=c)
        if cell.value is not None:
            cell.fill = FILL_NONE

# Column widths — keep original + coverage col
widths = {"A": 12, "B": 45, "C": 8, "D": 12, "E": 10, "F": 22, "G": 42, "H": 55, "I": 60}
for col, w in widths.items():
    ws2.column_dimensions[col].width = w

ws2.freeze_panes = "A2"
ws2.auto_filter.ref = ws2.dimensions

# ============================================================================
# Tab 3 — Scope Creep — CR Themes
# ============================================================================
ws3 = wb["3. Scope Creep — CR Themes"]

ws3["A1"].font = TITLE
ws3["A1"].alignment = Alignment(horizontal="left", vertical="center")
ws3.row_dimensions[1].height = 22
ws3["A2"].font = SUBTITLE

# Header row — usually row 3
for c in range(1, ws3.max_column + 1):
    cell = ws3.cell(row=3, column=c)
    if cell.value:
        style_table_header(cell)
ws3.row_dimensions[3].height = 32

# Body rows — themes
for r in range(4, ws3.max_row + 1):
    theme = str(ws3.cell(row=r, column=1).value or "")
    if theme.upper() == "NOTES" or theme.startswith(tuple(f"{i}." for i in range(1, 10))):
        # NOTES section — style as plain body
        for c in range(1, ws3.max_column + 1):
            cell = ws3.cell(row=r, column=c)
            if cell.value is not None:
                cell.font = BODY
                cell.alignment = WRAP_TOP
                cell.border = BORDER
        if theme.upper() == "NOTES":
            style_section_header(ws3.cell(row=r, column=1))
            ws3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ws3.max_column)
        continue

    # Determine row fill
    row_fill = FILL_CREEP  # default for scope creep themes
    if "RAISED" in theme:
        row_fill = FILL_CR_RAISED
    elif "NEW RECOMMENDED" in theme:
        row_fill = FILL_CR_NEW

    for c in range(1, ws3.max_column + 1):
        cell = ws3.cell(row=r, column=c)
        if cell.value is None:
            continue
        cell.font = BODY
        cell.alignment = WRAP_TOP
        cell.border = BORDER
        cell.fill = row_fill

# Column widths
widths3 = {"A": 34, "B": 36, "C": 8, "D": 14, "E": 14, "F": 22, "G": 50, "H": 60, "I": 42}
for col, w in widths3.items():
    ws3.column_dimensions[col].width = w

ws3.freeze_panes = "A4"
ws3.auto_filter.ref = ws3.dimensions

# ============================================================================
# Tab 4 — Borderline
# ============================================================================
ws4 = wb["4. Borderline"]

ws4["A1"].font = TITLE
ws4["A1"].alignment = Alignment(horizontal="left", vertical="center")
ws4.row_dimensions[1].height = 22
ws4["A2"].font = SUBTITLE

# Header row
header_row = None
for r in range(1, 10):
    if str(ws4.cell(row=r, column=1).value or "").strip() == "Ticket":
        header_row = r
        break
if header_row:
    for c in range(1, ws4.max_column + 1):
        cell = ws4.cell(row=header_row, column=c)
        if cell.value:
            style_table_header(cell)
    ws4.row_dimensions[header_row].height = 32

# Body rows — all borderline
for r in range(header_row + 1 if header_row else 4, ws4.max_row + 1):
    val = str(ws4.cell(row=r, column=1).value or "")
    if val in ("RESOLUTION BUCKETS",) or val == "Bucket":
        if val == "RESOLUTION BUCKETS":
            style_section_header(ws4.cell(row=r, column=1))
            ws4.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ws4.max_column)
        elif val == "Bucket":
            for c in range(1, ws4.max_column + 1):
                cell = ws4.cell(row=r, column=c)
                if cell.value:
                    style_table_header(cell)
        continue

    # Style body + apply fill
    approach = str(ws4.cell(row=r, column=6).value or "")
    row_fill = FILL_BORDER
    if "COVERED BY" in approach:
        row_fill = FILL_CR_COVERED
    elif "Goodwill" in approach or "Absorb" in approach:
        row_fill = FILL_BORDER
    elif "→ CR-" in approach:
        row_fill = FILL_CR_RAISED if ("CR01" in approach or "CR02" in approach) else FILL_BORDER

    for c in range(1, ws4.max_column + 1):
        cell = ws4.cell(row=r, column=c)
        if cell.value is None:
            continue
        cell.font = BODY
        cell.alignment = WRAP_TOP
        cell.border = BORDER
        cell.fill = row_fill

# Column widths
widths4 = {"A": 12, "B": 45, "C": 10, "D": 14, "E": 10, "F": 42, "G": 40, "H": 55, "I": 42}
for col, w in widths4.items():
    ws4.column_dimensions[col].width = w

ws4.freeze_panes = f"A{(header_row or 3) + 1}"
ws4.auto_filter.ref = ws4.dimensions

# ============================================================================
# Tab 5 — Review Findings
# ============================================================================
ws5 = wb["5. Review Findings"]

ws5["A1"].font = TITLE
ws5["A1"].alignment = Alignment(horizontal="left", vertical="center")
ws5.row_dimensions[1].height = 22
ws5["A2"].font = SUBTITLE

# Known section banners (original + new) — detect and style
banner_keywords = {
    "METHODOLOGY": True,
    "DEFENSIBILITY SUMMARY": True,
    "NEGOTIATION STRATEGY": True,
    "CONCESSION STRATEGY": True,
    "SOURCE DOCUMENTS": True,
    "CR-AWARE NEGOTIATION STRATEGY": True,
}

for r in range(1, ws5.max_row + 1):
    val = str(ws5.cell(row=r, column=1).value or "").strip()
    # Section banners
    if val in banner_keywords or any(val.startswith(k) for k in banner_keywords):
        style_section_header(ws5.cell(row=r, column=1))
        # Merge banner across used columns (up to D)
        try:
            ws5.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        except Exception:
            pass
        continue

    # Table headers (single-word labels in bold-white context)
    if val in ("Step", "Tier", "#", "Type", "Document"):
        for c in range(1, ws5.max_column + 1):
            cell = ws5.cell(row=r, column=c)
            if cell.value:
                style_table_header(cell)
        continue

    # Body rows
    for c in range(1, ws5.max_column + 1):
        cell = ws5.cell(row=r, column=c)
        if cell.value is None:
            continue
        cell.font = BODY
        cell.alignment = WRAP_TOP
        cell.border = BORDER

# New CR-aware methodology rows 5–6 keep CR_NEW highlight
for r in range(1, ws5.max_row + 1):
    val = str(ws5.cell(row=r, column=1).value or "")
    if val in ("5. CR Discovery (17/04/26)", "6. Commercial Reconciliation"):
        for c in (1, 2):
            if ws5.cell(row=r, column=c).value is not None:
                ws5.cell(row=r, column=c).fill = FILL_CR_NEW

# Column widths
ws5.column_dimensions["A"].width = 32
ws5.column_dimensions["B"].width = 95
ws5.column_dimensions["C"].width = 22
ws5.column_dimensions["D"].width = 50

# Dynamic row heights for long text
for r in range(1, ws5.max_row + 1):
    has_long = any(
        ws5.cell(row=r, column=c).value and len(str(ws5.cell(row=r, column=c).value)) > 100
        for c in range(1, 5)
    )
    if has_long:
        ws5.row_dimensions[r].height = 48

# ============================================================================
wb.save(DST)
print(f"POLISHED {DST}")
